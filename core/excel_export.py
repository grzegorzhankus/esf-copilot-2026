from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from core.contracts import AnalysisResult


def export_to_excel(result: AnalysisResult) -> bytes:
    """
    Export analysis result to Excel with multiple sheets.

    Sheets:
    1. Summary - Overview of the analysis
    2. Base Metrics - Raw financial metrics extracted from XML
    3. KPIs - Calculated financial KPIs
    4. Red Flags - Detected warning signs
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create all sheets
    _create_summary_sheet(wb, result)
    _create_base_metrics_sheet(wb, result)
    _create_kpis_sheet(wb, result)
    _create_red_flags_sheet(wb, result)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _create_summary_sheet(wb: Workbook, result: AnalysisResult):
    """Create summary overview sheet."""
    ws = wb.create_sheet("Summary", 0)

    # Title
    ws['A1'] = "e-SF Copilot - Financial Analysis Summary"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells('A1:B1')

    # Metadata section
    row = 3
    ws[f'A{row}'] = "File Information"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    metadata_info = [
        ("Filename", result.metadata.filename),
        ("Schema ID", result.metadata.schema_id),
        ("Schema Slug", result.metadata.schema_id_slug),
        ("File Size (bytes)", result.metadata.file_size_bytes),
        ("Analyzed At", result.metadata.analyzed_at_utc),
    ]

    for label, value in metadata_info:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

    # Coverage section
    row += 1
    ws[f'A{row}'] = "Coverage"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    ws[f'A{row}'] = "Coverage Percentage"
    ws[f'B{row}'] = f"{result.coverage.percent}%"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    ws[f'A{row}'] = "Metrics Present"
    ws[f'B{row}'] = len(result.coverage.present)
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    ws[f'A{row}'] = "Metrics Missing"
    ws[f'B{row}'] = len(result.coverage.missing)
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    # Summary statistics
    if result.kpis:
        row += 1
        ws[f'A{row}'] = "KPI Summary"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        categories = {}
        for kpi in result.kpis:
            categories[kpi.category] = categories.get(kpi.category, 0) + 1

        for category, count in categories.items():
            ws[f'A{row}'] = category
            ws[f'B{row}'] = f"{count} KPIs"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

    # Red flags summary
    if result.red_flags:
        row += 1
        ws[f'A{row}'] = "Red Flags Summary"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        detected = [f for f in result.red_flags if f.detected]
        high = len([f for f in detected if f.severity == "high"])
        medium = len([f for f in detected if f.severity == "medium"])
        low = len([f for f in detected if f.severity == "low"])

        ws[f'A{row}'] = "Total Detected"
        ws[f'B{row}'] = len(detected)
        ws[f'A{row}'].font = Font(bold=True)
        if len(detected) > 0:
            ws[f'B{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        row += 1

        ws[f'A{row}'] = "High Severity"
        ws[f'B{row}'] = high
        if high > 0:
            ws[f'B{row}'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            ws[f'B{row}'].font = Font(color="FFFFFF", bold=True)
        row += 1

        ws[f'A{row}'] = "Medium Severity"
        ws[f'B{row}'] = medium
        if medium > 0:
            ws[f'B{row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        row += 1

        ws[f'A{row}'] = "Low Severity"
        ws[f'B{row}'] = low
        if low > 0:
            ws[f'B{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50


def _create_base_metrics_sheet(wb: Workbook, result: AnalysisResult):
    """Create base metrics sheet."""
    ws = wb.create_sheet("Base Metrics")

    # Create DataFrame
    data = []
    for m in result.metrics_base:
        data.append({
            "Key": m.key,
            "Value": m.value,
            "Unit": m.unit,
            "Source Reference": m.source_ref,
        })

    df = pd.DataFrame(data)

    # Write to sheet with header formatting
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Header row
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def _create_kpis_sheet(wb: Workbook, result: AnalysisResult):
    """Create KPIs sheet grouped by category."""
    ws = wb.create_sheet("KPIs")

    if not result.kpis:
        ws['A1'] = "No KPIs calculated"
        return

    # Create DataFrame
    data = []
    for kpi in result.kpis:
        data.append({
            "Category": kpi.category,
            "KPI Name": kpi.name,
            "Value": kpi.value if kpi.value is not None else "N/A",
            "Unit": kpi.unit,
            "Interpretation": kpi.interpretation,
        })

    df = pd.DataFrame(data)

    # Write to sheet with header formatting
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Header row
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width


def _create_red_flags_sheet(wb: Workbook, result: AnalysisResult):
    """Create red flags sheet."""
    ws = wb.create_sheet("Red Flags")

    if not result.red_flags:
        ws['A1'] = "No red flags detected"
        return

    # Create DataFrame
    data = []
    for flag in result.red_flags:
        data.append({
            "Status": "DETECTED" if flag.detected else "OK",
            "Severity": flag.severity.upper() if flag.detected else "-",
            "Title": flag.title,
            "Description": flag.description,
            "Details": flag.details,
        })

    df = pd.DataFrame(data)

    # Write to sheet with header formatting
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Header row
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            elif r_idx > 1:
                # Color code by status and severity
                status = ws.cell(row=r_idx, column=1).value
                severity = ws.cell(row=r_idx, column=2).value

                if status == "DETECTED":
                    if severity == "HIGH":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    elif severity == "MEDIUM":
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    elif severity == "LOW":
                        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 70)
        ws.column_dimensions[column_letter].width = adjusted_width
